from pathlib import Path
import re
import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / ".python-packages"))

import openpyxl
import xlrd


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = PROJECT_ROOT / "data" / "Industry_employmet&wages"
YEARS = ("2020", "2021")


def safe_sheet_name(name, used_names):
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", name or "Sheet")
    cleaned = cleaned[:31] or "Sheet"
    candidate = cleaned
    suffix = 1

    while candidate in used_names:
        suffix_text = f"_{suffix}"
        candidate = f"{cleaned[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1

    used_names.add(candidate)
    return candidate


def convert_workbook(source_path, target_path):
    source_book = xlrd.open_workbook(source_path)
    target_book = openpyxl.Workbook()
    target_book.remove(target_book.active)

    used_sheet_names = set()

    for source_sheet in source_book.sheets():
        target_sheet = target_book.create_sheet(
            title=safe_sheet_name(source_sheet.name, used_sheet_names)
        )

        for row_index in range(source_sheet.nrows):
            for col_index in range(source_sheet.ncols):
                value = source_sheet.cell_value(row_index, col_index)
                cell_type = source_sheet.cell_type(row_index, col_index)

                if cell_type == xlrd.XL_CELL_EMPTY:
                    value = None

                target_sheet.cell(
                    row=row_index + 1,
                    column=col_index + 1,
                    value=value,
                )

    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_book.save(target_path)


def main():
    converted = 0

    for year in YEARS:
        source_dir = DATA_ROOT / f"alloes{year}"
        target_dir = DATA_ROOT / f"all-oews-{year}"

        if not source_dir.exists():
            print(f"Skipping missing source folder: {source_dir}")
            continue

        for source_path in sorted(source_dir.glob("*.xls")):
            target_path = target_dir / f"{source_path.stem}.xlsx"
            convert_workbook(source_path, target_path)
            converted += 1
            print(f"Converted {source_path.relative_to(PROJECT_ROOT)} -> {target_path.relative_to(PROJECT_ROOT)}")

    print(f"Converted {converted} legacy OEWS workbooks.")


if __name__ == "__main__":
    main()
