from pathlib import Path
import csv
import json
import re
import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / ".python-packages"))

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = PROJECT_ROOT / "Data" / "Industry_employmet&wages"
HOUSEHOLD_LOOKUP_PATH = PROJECT_ROOT / "Data" / "household_burden" / "california_counties_burdened_households_series_ids.csv"
OUTPUT_PATH = DATA_ROOT / "oews_normalized.json"
LATEST_OUTPUT_PATH = DATA_ROOT / "oews_latest.json"
LATEST_COMPACT_OUTPUT_PATH = DATA_ROOT / "oews_latest_compact.json"
LATEST_YEAR_COUNT = 5

OEWS_REQUIRED_HEADERS = {
    "soc": "SOC Code",
    "occupation": "Occupational Title",
    "employment": "Employment Estimates",
    "annual_wage": "Mean Annual Wage",
}


def main():
    county_id_by_name = load_county_lookup(HOUSEHOLD_LOOKUP_PATH)
    years = get_latest_oews_years(DATA_ROOT, LATEST_YEAR_COUNT)
    rows = []

    print(f"Preprocessing OEWS years: {', '.join(str(year) for year in years)}")

    for year in years:
        folder = DATA_ROOT / f"all-oews-{year}"
        workbook_paths = [
            path
            for path in sorted(folder.glob("*.xlsx"))
            if "California Statewide" not in path.name and not path.name.startswith("cal$")
        ]

        print(f"{year}: parsing {len(workbook_paths)} regional workbooks")

        for workbook_path in workbook_paths:
            parsed_rows = parse_oews_workbook(workbook_path, year, county_id_by_name)
            rows.extend(parsed_rows)
            print(f"  {workbook_path.name}: {len(parsed_rows)} rows")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(rows, separators=(",", ":")), encoding="utf-8")

    latest_year = max(years)
    latest_rows = [row for row in rows if row["year"] == latest_year]
    LATEST_OUTPUT_PATH.write_text(json.dumps(latest_rows, separators=(",", ":")), encoding="utf-8")
    LATEST_COMPACT_OUTPUT_PATH.write_text(
        json.dumps(to_compact_oews(latest_rows, latest_year), separators=(",", ":")),
        encoding="utf-8",
    )

    print(f"Wrote {len(rows)} normalized OEWS rows to {OUTPUT_PATH.relative_to(PROJECT_ROOT)}")
    print(f"Wrote {len(latest_rows)} latest-year OEWS rows to {LATEST_OUTPUT_PATH.relative_to(PROJECT_ROOT)}")
    print(f"Wrote compact latest-year OEWS data to {LATEST_COMPACT_OUTPUT_PATH.relative_to(PROJECT_ROOT)}")


def to_compact_oews(rows, year):
    group_titles = {}
    compact_rows = []

    for row in rows:
        soc_code = row["soc_code"]

        if row["occupation_level"] != "group":
            continue

        group_titles[soc_code] = row["occupation_title"]

        compact_rows.append([
            row["county_id"],
            soc_code,
            row["employment"],
            row["mean_annual_wage"],
        ])

    return {
        "version": 1,
        "year": year,
        "columns": ["county_id", "soc_code", "employment", "mean_annual_wage"],
        "groups": sorted(group_titles.items()),
        "rows": compact_rows,
    }


def load_county_lookup(path):
    county_id_by_name = {}

    with path.open(newline="", encoding="utf-8-sig") as file:
        for row in csv.DictReader(file):
            county = clean_county_name(row.get("county"))
            county_id = clean_county_id(row.get("county_fips"))

            if county and county_id:
                county_id_by_name[normalize_county_name(county)] = county_id

    return county_id_by_name


def get_latest_oews_years(data_root, limit):
    years = []

    for folder in data_root.glob("all-oews-*"):
        match = re.search(r"all-oews-(\d{4})$", folder.name)

        if match:
            years.append(int(match.group(1)))

    return sorted(years)[-limit:]


def parse_oews_workbook(path, year, county_id_by_name):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[get_oews_sheet_name(workbook)]
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    counties = get_workbook_counties(rows)
    header_index = find_header_row(rows)

    if header_index is None:
      print(f"  WARNING: missing OEWS header row: {path.name}")
      return []

    if not counties:
      print(f"  WARNING: no county list, skipped: {path.name}")
      return []

    county_matches = []

    for county in counties:
        county_id = county_id_by_name.get(normalize_county_name(county))

        if county_id:
            county_matches.append({"county": county, "county_id": county_id})
        else:
            print(f"  WARNING: unmatched county {county!r} in {path.name}")

    if not county_matches:
        print(f"  WARNING: no matched counties, skipped: {path.name}")
        return []

    headers = [str(value or "").strip() for value in rows[header_index]]
    indexes = {
        "soc": find_exact_header_index(headers, OEWS_REQUIRED_HEADERS["soc"]),
        "occupation": find_exact_header_index(headers, OEWS_REQUIRED_HEADERS["occupation"]),
        "employment": find_header_index(headers, OEWS_REQUIRED_HEADERS["employment"]),
        "annual_wage": find_header_index(headers, OEWS_REQUIRED_HEADERS["annual_wage"]),
    }

    if any(index is None for index in indexes.values()):
        print(f"  WARNING: unexpected columns, skipped: {path.name}")
        print(f"  Headers: {headers}")
        return []

    occupation_rows = [
        row
        for row in (
            to_occupation_row(row, indexes)
            for row in rows[header_index + 1:]
        )
        if row
    ]
    group_rows = [
        row
        for row in occupation_rows
        if get_occupation_level(row["soc_code"]) == "group"
    ]
    normalized_rows = []
    source_file = str(path.relative_to(PROJECT_ROOT))

    for row in group_rows:
        for county in county_matches:
            normalized_rows.append({
                "year": year,
                "county": county["county"],
                "county_id": county["county_id"],
                "soc_code": row["soc_code"],
                "occupation_title": row["occupation_title"],
                "occupation_group": row["occupation_title"],
                "employment": row["employment"],
                "mean_annual_wage": row["mean_annual_wage"],
                "source_file": source_file,
                "occupation_level": "group",
            })

    return normalized_rows


def get_oews_sheet_name(workbook):
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip() == "OEWS Data":
            return sheet_name

    for sheet_name in workbook.sheetnames:
        if re.search(r"oews?data", sheet_name, re.IGNORECASE):
            return sheet_name

    return workbook.sheetnames[0]


def get_workbook_counties(rows):
    county_text = ""

    for row in rows:
        first_cell = str(row[0] or "") if row else ""

        if first_cell.startswith("Counties:"):
            county_text = first_cell
            break

    if not county_text:
        return []

    county_text = re.sub(r"^Counties:\s*", "", county_text, flags=re.IGNORECASE)
    county_text = re.sub(r"\s*,?\s+and\s+", ", ", county_text, flags=re.IGNORECASE)

    return [
        clean_county_name(county)
        for county in re.split(r"\s*,\s*", county_text)
        if clean_county_name(county)
    ]


def find_header_row(rows):
    for index, row in enumerate(rows):
        values = set(str(value or "").strip() for value in row)

        if OEWS_REQUIRED_HEADERS["soc"] in values and OEWS_REQUIRED_HEADERS["occupation"] in values:
            return index

    return None


def find_exact_header_index(headers, text):
    try:
        return headers.index(text)
    except ValueError:
        return None


def find_header_index(headers, text):
    text = text.lower()

    for index, header in enumerate(headers):
        if text in header.lower():
            return index

    return None


def to_occupation_row(row, indexes):
    soc_code = str(row[indexes["soc"]] or "").strip()
    occupation_title = str(row[indexes["occupation"]] or "").strip()

    if not is_soc_code(soc_code) or not occupation_title:
        return None

    return {
        "soc_code": soc_code,
        "occupation_title": occupation_title,
        "employment": parse_number(row[indexes["employment"]]),
        "mean_annual_wage": parse_number(row[indexes["annual_wage"]]),
    }


def get_group_title_lookup(rows):
    group_title_by_prefix = {}

    for row in rows:
        if row["soc_code"].endswith("-0000") and row["soc_code"] != "00-0000":
            group_title_by_prefix[get_soc_prefix(row["soc_code"])] = row["occupation_title"]

    return group_title_by_prefix


def get_occupation_group(row, group_title_by_prefix):
    if row["soc_code"] == "00-0000":
        return "Total all occupations"

    return group_title_by_prefix.get(get_soc_prefix(row["soc_code"]), row["occupation_title"])


def get_occupation_level(soc_code):
    if soc_code == "00-0000":
        return "total"

    if soc_code.endswith("-0000"):
        return "group"

    return "occupation"


def get_soc_prefix(soc_code):
    return str(soc_code)[:2]


def get_soc_group_code(soc_code):
    if not is_soc_code(soc_code):
        return None

    prefix = get_soc_prefix(soc_code)

    if prefix == "00":
        return None

    return f"{prefix}-0000"


def is_soc_code(value):
    return bool(re.fullmatch(r"\d{2}-\d{4}", str(value or "")))


def clean_county_name(value):
    county_name = re.sub(r",\s*California$", "", str(value or ""), flags=re.IGNORECASE)
    county_name = re.sub(r"^and\s+", "", county_name, flags=re.IGNORECASE).strip()

    if not county_name:
        return ""

    if re.search(r"\bcounty$", county_name, flags=re.IGNORECASE):
        return county_name

    return f"{county_name} County"


def normalize_county_name(value):
    county = clean_county_name(value)
    county = re.sub(r"\s+county$", "", county, flags=re.IGNORECASE)
    return re.sub(r"[^a-z0-9]", "", county.lower())


def clean_county_id(value):
    if value is None or value == "":
        return None

    digits = re.sub(r"\D", "", str(value))

    if len(digits) >= 5:
        return digits[-5:]

    return digits.zfill(5)


def parse_number(value):
    if value is None or value == "" or value == "-" or value == "**":
        return None

    if isinstance(value, (int, float)):
        return value

    cleaned = re.sub(r"[$,%\s,]", "", str(value))

    try:
        return float(cleaned)
    except ValueError:
        return None


if __name__ == "__main__":
    main()
